# -*- coding: utf-8 -*-
"""
Aplikasi Input RPS (Rencana Pembelajaran Semester) — UNSIA

Cara menjalankan:
    pip install -r requirements.txt
    streamlit run app.py

Struktur data master:
    data/<Nama Prodi>.xlsx   -> setiap file mewakili 1 Program Studi
        sheet 'Mata Kuliah'  -> kolom: No, Nama Mata Kuliah, Kode MK, SKS, Semester, Ranah Topik, Dosen Pengembang
        sheet 'CPL'          -> kolom: Kode CPL, Deskripsi CPL

Catatan teknis (widget perlu 2x klik):
    Semua widget interaktif diberi `key=` yang STABIL selama Mata Kuliah yang sama
    dipilih (di-scope lewat helper `mk_key()`), dan berubah otomatis saat pindah MK.
    Ini mencegah Streamlit membuat ulang identitas widget di setiap rerun (penyebab
    umum gejala "perlu klik 2x"), sekaligus mencegah nilai lama "bocor" saat pindah MK.
"""

import io
import glob
import os
import json

import pandas as pd
import streamlit as st
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from google import genai
    from google.genai import types as genai_types
    GENAI_AVAILABLE = True
except ImportError:
    GENAI_AVAILABLE = False

from pdf_export import (
    build_pdf, with_code, CATATAN_POINTS, SKS_ROWS, BLOOM_TABLE, METODE_TABLE,
    BENTUK_TABLE, KOMPONEN_PENJELASAN, RUBRIK_ROWS,
)

# --------------------------------------------------------------------------
# Konstanta
# --------------------------------------------------------------------------
DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
DEFAULT_GEMINI_CONFIG_PATH = os.path.join(os.path.dirname(__file__), "config", "gemini_default.txt")


def load_default_gemini_config():
    """Baca config/gemini_default.txt -> {'api_key': ..., 'model': ...} (string kosong kalau tidak ada)."""
    result = {"api_key": "", "model": "gemini-2.0-flash"}
    if not os.path.exists(DEFAULT_GEMINI_CONFIG_PATH):
        return result
    try:
        with open(DEFAULT_GEMINI_CONFIG_PATH, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, _, value = line.partition("=")
                key, value = key.strip(), value.strip()
                if key == "GEMINI_API_KEY" and value:
                    result["api_key"] = value
                elif key == "GEMINI_MODEL" and value:
                    result["model"] = value
    except Exception:
        pass
    return result

BLOOM_LEVELS = ["C1", "C2", "C3", "C4", "C5", "C6"]
METODE_OPTIONS = ["SGD", "RPlS", "DL", "SDL", "CoL", "CbL", "CtL", "PjBL", "PBL"]
BENTUK_OPTIONS = ["EL-1", "EL-2", "EL-3", "EL-4", "EL-5", "EL-6", "EL-7", "EL-8"]

BLOOM_INFO = {
    "C1": "Remembering", "C2": "Understanding", "C3": "Applying",
    "C4": "Analyzing", "C5": "Evaluating", "C6": "Creating",
}
METODE_INFO = {
    "SGD": "Small Group Discussion", "RPlS": "Role-Play & Simulation",
    "DL": "Discovery Learning", "SDL": "Self-Directed Learning",
    "CoL": "Cooperative Learning", "CbL": "Collaborative Learning",
    "CtL": "Contextual Learning", "PjBL": "Problem Based Learning & Inquiry",
    "PBL": "Project Based Learning",
}
BENTUK_INFO = {
    "EL-1": "Video E-Learning", "EL-2": "Discussion at Forum",
    "EL-3": "Reading Module", "EL-4": "Video Conference / Webinar",
    "EL-5": "E-simulation using software (Virtual Lab)",
    "EL-6": "E-learning Link (jurnal/pustaka online)",
    "EL-7": "Vlog Presentation", "EL-8": "Assignment",
}


def info_tooltip(info_dict):
    return "  \n".join(f"**{k}** — {v}" for k, v in info_dict.items())


N_CPL_WAJIB = 5
N_MINGGU = 16
KATEGORI_PENILAIAN = ["Kehadiran dan Sikap", "Tugas", "UTS", "UAS"]
BOBOT_KATEGORI = {"Kehadiran dan Sikap": 30, "Tugas": 20, "UTS": 20, "UAS": 30}
HEADER_BLUE = "#B7DDE8"  # warna asli header tabel RPS UNSIA

st.set_page_config(page_title="RPS Builder — UNSIA", layout="wide")


# --------------------------------------------------------------------------
# Asisten AI (Google Gemini) — opsional, menyarankan draf isi 1 pertemuan
# --------------------------------------------------------------------------
def get_ai_suggestion(nama_mk, deskripsi_mk, cpmk_desc, sub_cpmk_desc, minggu, api_key, model_name="gemini-2.0-flash"):
    client = genai.Client(api_key=api_key)
    prompt = f"""Anda membantu dosen menyusun RPS (Rencana Pembelajaran Semester) untuk mata kuliah "{nama_mk}".

Deskripsi mata kuliah ini secara keseluruhan: {deskripsi_mk or "(belum diisi)"}

Konteks pertemuan minggu ke-{minggu} dari 16 minggu perkuliahan:
- CPMK terkait: {cpmk_desc or "(belum diisi)"}
- Sub-CPMK (jika sudah diisi dosen): {sub_cpmk_desc or "(belum diisi)"}

Gunakan deskripsi mata kuliah dan CPMK di atas sebagai konteks utama supaya saran yang diberikan
relevan dan konsisten dengan arah keseluruhan mata kuliah, bukan generik.

Sarankan draf singkat, konkret, dan realistis untuk SATU pertemuan ini saja (bukan seluruh semester),
dalam Bahasa Indonesia:
- materi: poin-poin utama materi pembelajaran minggu ini (boleh berupa daftar singkat)
- tugas: deskripsi tugas/quiz/assignment yang sesuai untuk pertemuan ini
- kriteria: kriteria penilaian untuk tugas tersebut
- indikator: indikator penilaian yang terukur"""

    config = genai_types.GenerateContentConfig(
        response_mime_type="application/json",
        response_schema={
            "type": "object",
            "properties": {
                "materi": {"type": "string"},
                "tugas": {"type": "string"},
                "kriteria": {"type": "string"},
                "indikator": {"type": "string"},
            },
            "required": ["materi", "tugas", "kriteria", "indikator"],
        },
    )
    response = client.models.generate_content(model=model_name, contents=prompt, config=config)
    return json.loads(response.text)



# --------------------------------------------------------------------------
# Data master
# --------------------------------------------------------------------------
@st.cache_data
def list_prodi():
    files = glob.glob(os.path.join(DATA_DIR, "*.xlsx"))
    return sorted([os.path.splitext(os.path.basename(f))[0] for f in files])


@st.cache_data
def load_master(prodi_name):
    path = os.path.join(DATA_DIR, f"{prodi_name}.xlsx")
    mk_df = pd.read_excel(path, sheet_name="Mata Kuliah")
    cpl_df = pd.read_excel(path, sheet_name="CPL")
    return mk_df, cpl_df


# --------------------------------------------------------------------------
# Session state
# --------------------------------------------------------------------------
def default_pertemuan():
    return {
        m: {
            "sub_cpmk_desc": "", "cpmk_ref": None, "bloom": [], "materi": "",
            "metode": [], "bentuk": [], "tugas": "", "kriteria": "", "indikator": "",
            "referensi": "", "bobot": 0,
        } for m in range(1, N_MINGGU + 1)
    }


def init_state():
    defaults = {
        "prodi_sel": None,
        "mk_sel": None,
        "cpl_selected": [],
        "info_umum": {
            "dosen_koordinator": "", "dosen_pengampu": "", "deskripsi_mk": "", "media": "", "modus": "",
            "nama_kaprodi": "", "nama_koordinator": "", "nama_penyusun": "",
            "nama_biro_pjm": "", "tanggal_dokumen": "",
        },
        "cpmk_data": {i: {"cpl_kode": None, "deskripsi": ""} for i in range(1, 6)},
        "pertemuan_data": default_pertemuan(),
        "komponen_data": {i: None for i in range(1, 6)},
        "referensi_data": [],
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def reset_rps_state():
    """Reset semua isian RPS (dipanggil saat ganti Mata Kuliah)."""
    keys = ["cpl_selected", "info_umum", "cpmk_data", "pertemuan_data",
            "komponen_data", "referensi_data"]
    for k in keys:
        if k in st.session_state:
            del st.session_state[k]
    init_state()


# --------------------------------------------------------------------------
# Konversi Pertemuan <-> DataFrame (dipakai untuk template impor/ekspor progres)
# --------------------------------------------------------------------------
def pertemuan_to_df(pertemuan_data):
    rows = []
    for m in range(1, N_MINGGU + 1):
        p = pertemuan_data[m]
        rows.append({
            "Minggu": m,
            "Sub-CPMK": p.get("sub_cpmk_desc", ""),
            "CPMK Ref": p.get("cpmk_ref") or "-",
            "Bloom": ", ".join(p.get("bloom", [])),
            "Materi": p.get("materi", ""),
            "Metode": ", ".join(p.get("metode", [])),
            "Bentuk Online": ", ".join(p.get("bentuk", [])),
            "Deskripsi Tugas": p.get("tugas", ""),
            "Kriteria": p.get("kriteria", ""),
            "Indikator": p.get("indikator", ""),
            "Referensi": p.get("referensi", ""),
            "Bobot (%)": p.get("bobot", 0),
        })
    return pd.DataFrame(rows)


def df_to_pertemuan(df):
    data = {}
    for _, row in df.iterrows():
        m = int(row["Minggu"])
        cpmk_ref = row["CPMK Ref"]
        data[m] = {
            "sub_cpmk_desc": row["Sub-CPMK"] or "",
            "cpmk_ref": None if cpmk_ref in ("-", None, "") else cpmk_ref,
            "bloom": [x.strip() for x in str(row["Bloom"] or "").split(",") if x.strip()],
            "materi": row["Materi"] or "",
            "metode": [x.strip() for x in str(row["Metode"] or "").split(",") if x.strip()],
            "bentuk": [x.strip() for x in str(row["Bentuk Online"] or "").split(",") if x.strip()],
            "tugas": row["Deskripsi Tugas"] or "",
            "kriteria": row["Kriteria"] or "",
            "indikator": row["Indikator"] or "",
            "referensi": row["Referensi"] or "",
            "bobot": int(row["Bobot (%)"]) if pd.notna(row["Bobot (%)"]) else 0,
        }
    return data


# --------------------------------------------------------------------------
# Simpan / muat progres — format Excel (bukan JSON, supaya bisa dibuka & dicek
# manual oleh dosen di Excel biasa)
# --------------------------------------------------------------------------
def serialize_progress_excel(mk_row):
    wb = Workbook()

    ws_meta = wb.active
    ws_meta.title = "Meta"
    ws_meta.append(["Key", "Value"])
    ws_meta.append(["prodi_sel", st.session_state.prodi_sel])
    ws_meta.append(["mk_sel", st.session_state.mk_sel])
    info = st.session_state.info_umum
    ws_meta.append(["dosen_pengampu", info["dosen_pengampu"]])
    ws_meta.append(["deskripsi_mk", info["deskripsi_mk"]])
    ws_meta.append(["media", info["media"]])
    ws_meta.append(["modus", info["modus"]])

    ws_cpl = wb.create_sheet("CPL_Selected")
    ws_cpl.append(["Kode CPL"])
    for k in st.session_state.cpl_selected:
        ws_cpl.append([k])

    ws_cpmk = wb.create_sheet("CPMK")
    ws_cpmk.append(["No", "Kode CPL", "Deskripsi"])
    for i in range(1, 6):
        c = st.session_state.cpmk_data[i]
        ws_cpmk.append([i, c["cpl_kode"], c["deskripsi"]])

    ws_komp = wb.create_sheet("Komponen")
    ws_komp.append(["CPMK", "Kategori"])
    for i in range(1, 6):
        ws_komp.append([f"CPMK-{i}", st.session_state.komponen_data.get(i) or ""])

    ws_ref = wb.create_sheet("Referensi")
    ws_ref.append(["No", "Sitasi"])
    for i, ref in enumerate(st.session_state.referensi_data, start=1):
        ws_ref.append([i, ref["sitasi"]])

    ws_prt = wb.create_sheet("Pertemuan")
    df = pertemuan_to_df(st.session_state.pertemuan_data)
    ws_prt.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws_prt.append(list(row))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def load_progress_excel(uploaded_file):
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)

    if "Meta" in wb.sheetnames:
        ws = wb["Meta"]
        meta = {row[0]: row[1] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}
        st.session_state.prodi_sel = meta.get("prodi_sel")
        st.session_state.mk_sel = meta.get("mk_sel")
        st.session_state.info_umum = {
            "dosen_pengampu": meta.get("dosen_pengampu") or "",
            "deskripsi_mk": meta.get("deskripsi_mk") or "",
            "media": meta.get("media") or "",
            "modus": meta.get("modus") or "",
        }

    if "CPL_Selected" in wb.sheetnames:
        ws = wb["CPL_Selected"]
        st.session_state.cpl_selected = [
            row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]
        ]

    if "CPMK" in wb.sheetnames:
        ws = wb["CPMK"]
        cpmk_data = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            no, cpl_kode, desk = row[0], row[1], row[2]
            if no:
                cpmk_data[int(no)] = {"cpl_kode": cpl_kode, "deskripsi": desk or ""}
        if cpmk_data:
            st.session_state.cpmk_data = cpmk_data

    if "Komponen" in wb.sheetnames:
        ws = wb["Komponen"]
        komponen_data = {}
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            komponen_data[i] = row[1] or None
        if komponen_data:
            st.session_state.komponen_data = komponen_data

    if "Referensi" in wb.sheetnames:
        ws = wb["Referensi"]
        st.session_state.referensi_data = [
            {"sitasi": row[1]} for row in ws.iter_rows(min_row=2, values_only=True) if row[1]
        ]

    if "Pertemuan" in wb.sheetnames:
        ws = wb["Pertemuan"]
        values = list(ws.iter_rows(min_row=1, values_only=True))
        if values:
            cols = values[0]
            df = pd.DataFrame(values[1:], columns=cols)
            st.session_state.pertemuan_data = df_to_pertemuan(df)


# --------------------------------------------------------------------------
# Export ke Excel (struktur menyatu — meniru dokumen RPS asli)
# --------------------------------------------------------------------------
def export_to_excel(mk_row, cpl_df):
    wb = Workbook()
    ws = wb.active
    ws.title = "RPS"

    header_fill = PatternFill("solid", fgColor="B7DDE8")
    bold = Font(bold=True)
    border = Border(*(Side(style="thin", color="D9D9D9"),) * 4)
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center")
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12, color="1E3A5F")

    info = st.session_state.info_umum
    cpmk_data = st.session_state.cpmk_data
    komponen_data = st.session_state.komponen_data

    def set_cell(row, col, value, fill=None, font=None, merge_to_col=None):
        c = ws.cell(row=row, column=col, value=value)
        c.border = border
        c.alignment = wrap
        if fill:
            c.fill = fill
        if font:
            c.font = font
        if merge_to_col:
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to_col)
            for cc in range(col + 1, merge_to_col + 1):
                ws.cell(row=row, column=cc).border = border
        return c

    r = 1
    ws.cell(row=r, column=1, value="RENCANA PEMBELAJARAN SEMESTER (RPS)").font = title_font
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 2

    # --- 4 baris identitas MK ---
    set_cell(r, 1, "Program Studi", fill=header_fill, font=bold)
    set_cell(r, 2, st.session_state.prodi_sel)
    set_cell(r, 3, "Semester", fill=header_fill, font=bold)
    set_cell(r, 4, mk_row["Semester"])
    r += 1
    set_cell(r, 1, "Mata Kuliah", fill=header_fill, font=bold)
    set_cell(r, 2, mk_row["Nama Mata Kuliah"])
    set_cell(r, 3, "Beban SKS", fill=header_fill, font=bold)
    set_cell(r, 4, f'{mk_row["SKS"]} SKS')
    r += 1
    set_cell(r, 1, "Ranah Topik", fill=header_fill, font=bold)
    set_cell(r, 2, mk_row.get("Ranah Topik", "-"))
    set_cell(r, 3, "Dosen Pengampu (Koordinator & Anggota)", fill=header_fill, font=bold)
    set_cell(r, 4, info["dosen_pengampu"])
    r += 1
    set_cell(r, 1, "Kode Mata Kuliah", fill=header_fill, font=bold)
    set_cell(r, 2, mk_row["Kode MK"], merge_to_col=4)
    r += 1

    # --- CPL ---
    cpl_selected = [cpmk_data[i]["cpl_kode"] for i in range(1, 6)]
    seen = set()
    cpl_rows = []
    for kode in cpl_selected:
        if kode in seen:
            continue
        seen.add(kode)
        desk = cpl_df.loc[cpl_df["Kode CPL"] == kode, "Deskripsi CPL"]
        cpl_rows.append((kode, desk.values[0] if len(desk) else ""))
    cpl_start = r
    for idx, (kode, desk) in enumerate(cpl_rows):
        label = "Capaian Pembelajaran Lulusan (CPL)" if idx == 0 else ""
        set_cell(r, 1, label, fill=header_fill, font=bold)
        set_cell(r, 2, kode)
        set_cell(r, 3, desk, merge_to_col=4)
        r += 1
    if len(cpl_rows) > 1:
        ws.merge_cells(start_row=cpl_start, start_column=1, end_row=cpl_start + len(cpl_rows) - 1, end_column=1)

    # --- CPMK ---
    cpmk_start = r
    for i in range(1, 6):
        c = cpmk_data[i]
        label = "Capaian Pembelajaran Mata Kuliah (CP-MK)" if i == 1 else ""
        set_cell(r, 1, label, fill=header_fill, font=bold)
        set_cell(r, 2, f"CPMK-{i}")
        set_cell(r, 3, with_code(c["deskripsi"], c["cpl_kode"]), merge_to_col=4)
        r += 1
    ws.merge_cells(start_row=cpmk_start, start_column=1, end_row=cpmk_start + 4, end_column=1)

    # --- Deskripsi Mata Kuliah ---
    set_cell(r, 1, "Deskripsi Mata Kuliah", fill=header_fill, font=bold)
    set_cell(r, 2, info["deskripsi_mk"], merge_to_col=4)
    r += 1

    # --- Komponen Penilaian: tabel ceklist (CPMK x kategori) + baris bobot ---
    komp_start = r
    kategori_list = list(BOBOT_KATEGORI.keys())
    n_kat = len(kategori_list)
    # header kategori di baris pertama komponen (kolom B dst, dalam batas 4 kolom
    # tabel utama kita satukan jadi teks agar tetap di dalam tabel utama)
    set_cell(r, 1, "Komponen Penilaian", fill=header_fill, font=bold)
    header_line = " | ".join(kategori_list)
    set_cell(r, 2, "CPMK", font=bold)
    set_cell(r, 3, header_line, font=bold, merge_to_col=4)
    r += 1
    for i in range(1, 6):
        cpmk_kat = komponen_data.get(i)
        marks = " | ".join("\u2713" if k == cpmk_kat else "\u00b7" for k in kategori_list)
        set_cell(r, 1, "")
        set_cell(r, 2, f"CPMK-{i}")
        c3 = set_cell(r, 3, marks, merge_to_col=4)
        c3.alignment = Alignment(horizontal="center", vertical="center")
        r += 1
    set_cell(r, 1, "")
    set_cell(r, 2, "Bobot", font=bold)
    bobot_line = " | ".join(f"{k} {v}%" for k, v in BOBOT_KATEGORI.items())
    set_cell(r, 3, bobot_line, font=bold, merge_to_col=4)
    r += 1
    ws.merge_cells(start_row=komp_start, start_column=1, end_row=r - 1, end_column=1)

    # --- Media & Modus Pembelajaran ---
    set_cell(r, 1, "Media Pembelajaran", fill=header_fill, font=bold)
    set_cell(r, 2, info["media"], merge_to_col=4)
    r += 1
    set_cell(r, 1, "Modus Pembelajaran", fill=header_fill, font=bold)
    set_cell(r, 2, info["modus"], merge_to_col=4)
    r += 2

    # --- Tabel 16 Pertemuan ---
    ws.cell(row=r, column=1, value="Rencana Pembelajaran per Minggu").font = section_font
    r += 1
    headers = ["Minggu", "Kemampuan Akhir (Sub-CPMK)", "Bloom's Taxonomy", "Materi Pembelajaran",
               "Metode Pembelajaran", "Bentuk Pembelajaran Online", "Deskripsi Quiz/Tugas/Assignment",
               "Kriteria Penilaian", "Indikator Penilaian", "Referensi", "Bobot Penilaian (%)"]
    for c_i, h in enumerate(headers, start=1):
        cc = ws.cell(row=r, column=c_i, value=h)
        cc.fill = header_fill
        cc.font = bold
        cc.border = border
        cc.alignment = wrap
    r += 1
    for m in range(1, N_MINGGU + 1):
        p = st.session_state.pertemuan_data[m]
        sub_text = with_code(p.get("sub_cpmk_desc", ""), p.get("cpmk_ref"))
        vals = [m, sub_text, ", ".join(p["bloom"]), p["materi"], ", ".join(p["metode"]),
                ", ".join(p["bentuk"]), p["tugas"], p["kriteria"], p["indikator"],
                p["referensi"], p["bobot"]]
        for c_i, v in enumerate(vals, start=1):
            cc = ws.cell(row=r, column=c_i, value=v)
            cc.border = border
            cc.alignment = wrap
        r += 1
    r += 1

    # --- Referensi ---
    ws.cell(row=r, column=1, value="Referensi").font = section_font
    r += 1
    for i, ref in enumerate(st.session_state.referensi_data, start=1):
        ws.cell(row=r, column=1, value=f"{i}. {ref['sitasi']}").alignment = wrap
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1
    r += 1

    # --- Catatan ---
    ws.cell(row=r, column=1, value="Catatan").font = section_font
    r += 1
    for i, point in enumerate(CATATAN_POINTS, start=1):
        ws.cell(row=r, column=1, value=f"{i}. {point}").alignment = wrap
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1
    r += 1

    # --- Pengertian 1 SKS ---
    ws.cell(row=r, column=1, value="Pengertian 1 SKS dalam Bentuk Pembelajaran").font = section_font
    r += 1
    set_cell(r, 1, "", fill=header_fill, merge_to_col=3)
    set_cell(r, 4, "Durasi (Jam)", fill=header_fill, font=bold)
    r += 1
    for kode, judul, detail, durasi in SKS_ROWS:
        set_cell(r, 1, f"{kode}. {judul} — {detail}", merge_to_col=3)
        set_cell(r, 4, durasi)
        r += 1
    r += 1

    # --- Legenda Bloom's / Metode / Bentuk ---
    ws.cell(row=r, column=1, value="Legenda Bloom's Taxonomy, Metode & Bentuk Pembelajaran").font = section_font
    r += 1
    set_cell(r, 1, "Bloom's Taxonomy", fill=header_fill, font=bold, merge_to_col=2)
    set_cell(r, 3, "Metode Pembelajaran SCL", fill=header_fill, font=bold)
    set_cell(r, 4, "Bentuk Pembelajaran On-Line", fill=header_fill, font=bold)
    r += 1
    max_legend_rows = max(len(BLOOM_TABLE), len(METODE_TABLE), len(BENTUK_TABLE))
    for i in range(max_legend_rows):
        bloom_txt = f"{BLOOM_TABLE[i][1]} ({BLOOM_TABLE[i][2]})" if i < len(BLOOM_TABLE) else ""
        metode_txt = f"{METODE_TABLE[i][1]} ({METODE_TABLE[i][2]})" if i < len(METODE_TABLE) else ""
        bentuk_txt = f"{BENTUK_TABLE[i][1]} ({BENTUK_TABLE[i][2]})" if i < len(BENTUK_TABLE) else ""
        set_cell(r, 1, bloom_txt, merge_to_col=2)
        set_cell(r, 3, metode_txt)
        set_cell(r, 4, bentuk_txt)
        r += 1
    r += 1

    # --- Penjelasan Komponen Penilaian ---
    ws.cell(row=r, column=1, value="Penjelasan Komponen Penilaian").font = section_font
    r += 1
    ws.cell(row=r, column=1,
            value="Proses penilaian pada mata kuliah ini dibedakan dalam 4 komponen, di antaranya:").alignment = wrap
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 1
    for label, teks in KOMPONEN_PENJELASAN:
        ws.cell(row=r, column=1, value=f"{label} — {teks}").alignment = wrap
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1
    r += 1

    # --- Rubrik Penilaian ---
    ws.cell(row=r, column=1, value="Rubrik Penilaian").font = section_font
    r += 1
    set_cell(r, 1, "Jenjang", fill=header_fill, font=bold)
    set_cell(r, 2, "Angka/Skor", fill=header_fill, font=bold)
    set_cell(r, 3, "Deskripsi/Indikator Kinerja", fill=header_fill, font=bold, merge_to_col=4)
    r += 1
    for jenjang, skor, desk in RUBRIK_ROWS:
        set_cell(r, 1, jenjang)
        set_cell(r, 2, skor)
        set_cell(r, 3, desk, merge_to_col=4)
        r += 1
    r += 2

    # --- Blok Validasi (nama diisi manual, tanpa QR) ---
    tgl = info.get("tanggal_dokumen") or "…………………"
    set_cell(r, 1, f"Disetujui,\nTgl: {tgl}", fill=header_fill, font=bold)
    set_cell(r, 2, f"Diperiksa,\nTgl: {tgl}", fill=header_fill, font=bold)
    set_cell(r, 3, f"Dibuat,\nTgl: {tgl}", fill=header_fill, font=bold, merge_to_col=4)
    r += 1
    set_cell(r, 1, "Ketua Prodi")
    set_cell(r, 2, "Koordinator Mata Kuliah/Bidang Keahlian")
    set_cell(r, 3, "Dosen yang bersangkutan", merge_to_col=4)
    r += 3
    set_cell(r, 1, info.get("nama_kaprodi") or "…………………………")
    set_cell(r, 2, info.get("nama_koordinator") or "…………………………")
    set_cell(r, 3, info.get("nama_penyusun") or "…………………………", merge_to_col=4)
    r += 2
    set_cell(r, 1, "Periksa: Biro Penjaminan Mutu", fill=header_fill, font=bold, merge_to_col=4)
    r += 3
    set_cell(r, 1, info.get("nama_biro_pjm") or "…………………………", merge_to_col=4)
    r += 1

    widths = [22, 18, 14, 34, 16, 18, 26, 20, 20, 10, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# --------------------------------------------------------------------------
# UI
# --------------------------------------------------------------------------
init_state()

st.title("📘 RPS Builder")
st.caption("Universitas Siber Asia")

with st.sidebar:
    st.header("Pilih Mata Kuliah")

    prodi_options = list_prodi()
    if not prodi_options:
        st.error("Tidak ada data Prodi di folder data/. Tambahkan file xlsx (lihat README).")
        st.stop()

    prodi_sel = st.selectbox(
        "Program Studi", prodi_options,
        index=prodi_options.index(st.session_state.prodi_sel) if st.session_state.prodi_sel in prodi_options else 0,
        key="prodi_selectbox",
    )
    if prodi_sel != st.session_state.prodi_sel:
        st.session_state.prodi_sel = prodi_sel
        st.session_state.mk_sel = None

    mk_df, cpl_df = load_master(prodi_sel)
    mk_options = mk_df["Nama Mata Kuliah"].tolist()
    mk_sel_name = st.selectbox(
        "Mata Kuliah", mk_options,
        index=mk_options.index(st.session_state.mk_sel) if st.session_state.mk_sel in mk_options else 0,
        key="mk_selectbox",
    )

    if mk_sel_name != st.session_state.mk_sel:
        reset_rps_state()
        st.session_state.prodi_sel = prodi_sel
        st.session_state.mk_sel = mk_sel_name
        st.rerun()

    mk_row = mk_df[mk_df["Nama Mata Kuliah"] == mk_sel_name].iloc[0]
    st.markdown(f"""
    **Kode MK:** {mk_row['Kode MK']}
    **SKS:** {mk_row['SKS']} · **Semester:** {mk_row['Semester']}
    """)

    st.divider()
    st.subheader("💾 Progres")
    st.caption("Format Excel — bisa dibuka & dicek manual bila perlu.")
    st.download_button(
        "Unduh Progres (.xlsx)", data=serialize_progress_excel(mk_row),
        file_name=f"progres_{mk_row['Kode MK']}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_progress_btn",
    )
    up = st.file_uploader("Muat Progres (.xlsx)", type=["xlsx"], key="progress_uploader")
    if up is not None:
        load_progress_excel(up)
        st.success("Progres dimuat. Silakan lanjutkan pengisian.")
        st.rerun()

    st.divider()
    st.subheader("🤖 Asisten AI (Gemini)")
    if not GENAI_AVAILABLE:
        st.caption(
            "Paket `google-genai` belum terpasang. Jalankan `pip install google-genai` "
            "(sudah ada di requirements.txt), lalu restart aplikasi."
        )
    else:
        default_cfg = load_default_gemini_config()
        has_default = bool(default_cfg["api_key"])

        source_options = (["Default (disediakan Prodi)"] if has_default else []) + ["API Key Sendiri"]
        source = st.radio("Sumber API Key", source_options, key="gemini_key_source")

        if source == "Default (disediakan Prodi)":
            st.caption(f"Memakai API key default yang disediakan. Model: `{default_cfg['model']}`")
            st.session_state["gemini_api_key_effective"] = default_cfg["api_key"]
            st.session_state["gemini_model_effective"] = default_cfg["model"]
        else:
            st.caption(
                "Kunci API hanya dipakai untuk sesi ini, tidak disimpan permanen ke mana pun."
            )
            st.text_input("Google Gemini API Key", type="password", key="gemini_api_key")
            st.text_input(
                "Model Gemini", value="gemini-2.0-flash", key="gemini_model_name",
                help=(
                    "Ganti kalau muncul error kuota (429 RESOURCE_EXHAUSTED) untuk model ini — "
                    "coba model lain, mis. 'gemini-2.5-flash-lite', 'gemini-1.5-flash', dsb. "
                    "Ketersediaan kuota gratis per model bisa berbeda-beda dan berubah dari waktu ke waktu."
                ),
            )
            st.session_state["gemini_api_key_effective"] = st.session_state.get("gemini_api_key", "")
            st.session_state["gemini_model_effective"] = st.session_state.get("gemini_model_name") or "gemini-2.0-flash"

        if not has_default:
            st.caption(
                "Belum ada API key default. Admin bisa isi `config/gemini_default.txt` untuk "
                "menyediakan API key bersama sebagai opsi tambahan bagi semua Dosen."
            )


def mk_key(suffix):
    """Bikin widget key yang unik & stabil per Mata Kuliah aktif.

    Kunci tetap SAMA selama MK yang sama masih dipilih (mencegah Streamlit
    membuat ulang identitas widget setiap rerun -> penyebab gejala 'klik 2x'),
    tapi otomatis BEDA saat pindah MK (mencegah nilai lama nyangkut/bocor)."""
    return f"{st.session_state.mk_sel}__{suffix}"


tab_info, tab_cpl, tab_pertemuan, tab_ref, tab_nilai, tab_export = st.tabs(
    ["Info Umum", "CPL & CPMK", "16 Pertemuan", "Referensi", "Komponen Penilaian", "Pratinjau & Ekspor"]
)

# --- Tab: Info Umum ---
with tab_info:
    st.subheader("Informasi Umum RPS")
    info = st.session_state.info_umum
    c1, c2 = st.columns(2)
    info["dosen_koordinator"] = c1.text_input(
        "Dosen Koordinator", info["dosen_koordinator"], key=mk_key("dosen_koordinator"),
    )
    info["dosen_pengampu"] = c2.text_input(
        "Dosen Pengampu (Anggota)", info["dosen_pengampu"], key=mk_key("dosen_pengampu"),
    )
    info["deskripsi_mk"] = st.text_area(
        "Deskripsi Mata Kuliah", info["deskripsi_mk"], key=mk_key("deskripsi_mk"),
    )
    col1, col2 = st.columns(2)
    info["media"] = col1.text_area("Media Pembelajaran", info["media"], height=100, key=mk_key("media"))
    info["modus"] = col2.text_area("Modus Pembelajaran", info["modus"], height=100, key=mk_key("modus"))

# --- Tab: CPL & CPMK ---
with tab_cpl:
    st.subheader("Pilih CPL Mata Kuliah")
    st.caption(f"Wajib memilih tepat {N_CPL_WAJIB} CPL untuk mata kuliah ini.")
    cpl_all = cpl_df["Kode CPL"].tolist()
    cpl_label = {row["Kode CPL"]: f"{row['Kode CPL']} — {row['Deskripsi CPL'][:70]}…"
                 for _, row in cpl_df.iterrows()}
    selected = st.multiselect(
        "CPL", cpl_all, default=st.session_state.cpl_selected,
        format_func=lambda k: cpl_label[k],
        key=mk_key("cpl_multiselect"),
        help=(
            "Urutan CPL yang dipilih menentukan pemetaan ke CPMK: CPL pertama yang dicentang "
            "otomatis jadi rujukan CPMK-1, kedua jadi CPMK-2, dst. Untuk mengubah pemetaan, "
            "hapus semua lalu pilih ulang sesuai urutan yang diinginkan."
        ),
    )
    st.session_state.cpl_selected = selected

    if len(selected) != N_CPL_WAJIB:
        st.warning(f"Sudah dipilih {len(selected)} dari {N_CPL_WAJIB} CPL yang diwajibkan.")
    else:
        st.success(f"{N_CPL_WAJIB} CPL sudah lengkap.")

        st.divider()
        st.subheader("CPMK")
        st.caption(
            "CPL untuk tiap CPMK ditentukan otomatis dari urutan CPL yang dipilih di atas "
            "(CPL ke-1 → CPMK-1, CPL ke-2 → CPMK-2, dst.) — dosen tinggal isi deskripsinya. "
            "Kode CPL otomatis ditambahkan dalam kurung di akhir deskripsi CPMK pada hasil ekspor."
        )
        for i in range(1, 6):
            st.session_state.cpmk_data[i]["cpl_kode"] = selected[i - 1]
            st.session_state.cpmk_data[i]["deskripsi"] = st.text_area(
                f"Deskripsi CPMK-{i}  (rujukan: {selected[i - 1]})",
                st.session_state.cpmk_data[i]["deskripsi"],
                key=mk_key(f"cpmk_desk_{i}"), height=70,
            )
        st.success("5 CPL sudah otomatis terpetakan satu-satu ke CPMK-1 s/d CPMK-5, tanpa duplikasi.")

# --- Tab: 16 Pertemuan (accordion per minggu + opsi impor tabel) ---
with tab_pertemuan:
    st.subheader("Rincian 16 Pertemuan")

    with st.expander("📥 Impor dari Tabel (opsional) — isi banyak minggu sekaligus"):
        st.caption(
            "Unduh templatnya, isi di Excel (lebih leluasa untuk isi banyak baris sekaligus), "
            "lalu unggah kembali untuk mengisi otomatis ke-16 pertemuan. Kolom **CPMK Ref** diisi "
            "'CPMK-1' s/d 'CPMK-5' (atau '-' untuk UTS/UAS); kolom Bloom/Metode/Bentuk Online "
            "dipisah koma kalau lebih dari satu."
        )
        template_buf = io.BytesIO()
        pertemuan_to_df(default_pertemuan()).to_excel(template_buf, index=False)
        template_buf.seek(0)
        c1, c2 = st.columns(2)
        c1.download_button(
            "Unduh Template Kosong (.xlsx)", data=template_buf,
            file_name="template_16_pertemuan.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=mk_key("download_template_pertemuan"),
        )
        upfile = c2.file_uploader("Unggah Tabel Terisi", type=["xlsx", "csv"], key=mk_key("pertemuan_import"))
        if upfile is not None:
            try:
                imported_df = pd.read_csv(upfile) if upfile.name.endswith(".csv") else pd.read_excel(upfile)
                st.session_state.pertemuan_data = df_to_pertemuan(imported_df)
                st.success("Berhasil mengimpor data 16 pertemuan.")
                st.rerun()
            except Exception as e:
                st.error(f"Gagal mengimpor: {e}")

    st.caption(
        "Sub-CPMK diisi langsung per minggu, sertakan CPMK mana yang dirujuk lewat **CPMK Ref** "
        "(kode CPMK otomatis ditambahkan dalam kurung di akhir kalimat Sub-CPMK pada hasil ekspor)."
    )
    cpmk_ref_options = ["-"] + [f"CPMK-{i}" for i in range(1, 6)]
    for m in range(1, N_MINGGU + 1):
        p = st.session_state.pertemuan_data[m]
        label = f"Minggu {m}" + (" — UTS" if m == 8 else " — UAS" if m == 16 else "")
        preview = p["sub_cpmk_desc"] or p["materi"] or "belum diisi"
        with st.expander(f"{label} — {preview[:60]}{'…' if len(preview) > 60 else ''}"):
            c1, c2 = st.columns([3, 1])
            p["sub_cpmk_desc"] = c1.text_area("Sub-CPMK", p["sub_cpmk_desc"], key=mk_key(f"prt_sub_{m}"), height=70)
            current_ref = p["cpmk_ref"] or "-"
            ref_choice = c2.selectbox(
                "CPMK Ref", cpmk_ref_options,
                index=cpmk_ref_options.index(current_ref) if current_ref in cpmk_ref_options else 0,
                key=mk_key(f"prt_cpmkref_{m}"),
            )
            p["cpmk_ref"] = None if ref_choice == "-" else ref_choice

            if GENAI_AVAILABLE:
                if st.button("✨ Sarankan dengan AI", key=mk_key(f"ai_btn_{m}")):
                    api_key = st.session_state.get("gemini_api_key_effective")
                    if not api_key:
                        st.warning("Isi dulu Google Gemini API Key di sidebar (atau pilih sumber Default kalau tersedia).")
                    else:
                        with st.spinner("Meminta saran dari AI..."):
                            try:
                                cpmk_desc = ""
                                if p["cpmk_ref"]:
                                    idx_cpmk = int(p["cpmk_ref"].split("-")[1])
                                    cpmk_desc = st.session_state.cpmk_data[idx_cpmk]["deskripsi"]
                                model_name = st.session_state.get("gemini_model_effective") or "gemini-2.0-flash"
                                suggestion = get_ai_suggestion(
                                    mk_row["Nama Mata Kuliah"], st.session_state.info_umum.get("deskripsi_mk", ""),
                                    cpmk_desc, p["sub_cpmk_desc"], m, api_key,
                                    model_name=model_name,
                                )
                                st.session_state[mk_key(f"ai_suggestion_{m}")] = suggestion
                            except Exception as e:
                                msg = str(e)
                                if "RESOURCE_EXHAUSTED" in msg or "429" in msg:
                                    st.error(
                                        "Kuota Gemini untuk model ini habis/nol pada paket akun Bapak "
                                        f"(model: `{model_name}`). Ini bukan error dari aplikasi, melainkan "
                                        "dari sisi akun Google. Coba salah satu:\n\n"
                                        "1. Ganti **Model Gemini** di sidebar ke model lain (mis. "
                                        "`gemini-2.5-flash-lite` atau `gemini-1.5-flash`) — kuota gratis "
                                        "per model berbeda-beda.\n"
                                        "2. Aktifkan billing (pay-as-you-go) di [Google AI Studio]"
                                        "(https://aistudio.google.com) — biaya pemakaian ringan seperti ini "
                                        "biasanya sangat kecil.\n"
                                        "3. Tunggu beberapa saat lalu coba lagi (kuota harian/menit bisa reset)."
                                    )
                                else:
                                    st.error(f"Gagal meminta saran AI: {e}")

                suggestion = st.session_state.get(mk_key(f"ai_suggestion_{m}"))
                if suggestion:
                    st.info(
                        "**Saran AI** (tinjau dulu, belum diterapkan ke field di bawah):\n\n"
                        f"**Materi:** {suggestion.get('materi', '-')}\n\n"
                        f"**Tugas:** {suggestion.get('tugas', '-')}\n\n"
                        f"**Kriteria:** {suggestion.get('kriteria', '-')}\n\n"
                        f"**Indikator:** {suggestion.get('indikator', '-')}"
                    )
                    ca, cb = st.columns(2)
                    if ca.button("✅ Terapkan Saran", key=mk_key(f"ai_apply_{m}")):
                        new_materi = suggestion.get("materi", p["materi"])
                        new_tugas = suggestion.get("tugas", p["tugas"])
                        new_kriteria = suggestion.get("kriteria", p["kriteria"])
                        new_indikator = suggestion.get("indikator", p["indikator"])
                        # Tulis ke data kita sendiri...
                        p["materi"] = new_materi
                        p["tugas"] = new_tugas
                        p["kriteria"] = new_kriteria
                        p["indikator"] = new_indikator
                        # ...DAN ke key widget-nya langsung — widget dengan `key` tetap akan
                        # selalu memakai nilai dari session_state[key], bukan argumen `value=`
                        # yang kita berikan, begitu key tersebut pernah dibuat. Tanpa baris ini
                        # perubahan tidak akan pernah muncul di kotak teksnya.
                        st.session_state[mk_key(f"prt_materi_{m}")] = new_materi
                        st.session_state[mk_key(f"prt_tugas_{m}")] = new_tugas
                        st.session_state[mk_key(f"prt_kriteria_{m}")] = new_kriteria
                        st.session_state[mk_key(f"prt_indikator_{m}")] = new_indikator
                        del st.session_state[mk_key(f"ai_suggestion_{m}")]
                        st.rerun()
                    if cb.button("✖ Abaikan", key=mk_key(f"ai_dismiss_{m}")):
                        del st.session_state[mk_key(f"ai_suggestion_{m}")]
                        st.rerun()

            p["bloom"] = st.multiselect(
                "Bloom's Taxonomy Level", BLOOM_LEVELS, default=p["bloom"], key=mk_key(f"prt_bloom_{m}"),
                help=info_tooltip(BLOOM_INFO),
            )
            p["materi"] = st.text_area("Materi Pembelajaran", p["materi"], key=mk_key(f"prt_materi_{m}"))
            c3, c4 = st.columns(2)
            p["metode"] = c3.multiselect(
                "Metode Pembelajaran", METODE_OPTIONS, default=p["metode"], key=mk_key(f"prt_metode_{m}"),
                help=info_tooltip(METODE_INFO),
            )
            p["bentuk"] = c4.multiselect(
                "Bentuk Pembelajaran Online", BENTUK_OPTIONS, default=p["bentuk"], key=mk_key(f"prt_bentuk_{m}"),
                help=info_tooltip(BENTUK_INFO),
            )
            p["tugas"] = st.text_area("Deskripsi Tugas/Quiz/Assignment", p["tugas"], key=mk_key(f"prt_tugas_{m}"))
            c5, c6 = st.columns(2)
            p["kriteria"] = c5.text_area("Kriteria Penilaian", p["kriteria"], key=mk_key(f"prt_kriteria_{m}"))
            p["indikator"] = c6.text_area("Indikator Penilaian", p["indikator"], key=mk_key(f"prt_indikator_{m}"))
            c7, c8 = st.columns(2)
            p["referensi"] = c7.text_input("Referensi (nomor, pisah koma)", p["referensi"], key=mk_key(f"prt_ref_{m}"))
            p["bobot"] = c8.number_input("Bobot Penilaian (%)", 0, 100, p["bobot"], key=mk_key(f"prt_bobot_{m}"))

# --- Tab: Referensi ---
with tab_ref:
    st.subheader("Daftar Referensi")
    for idx, ref in enumerate(st.session_state.referensi_data):
        c1, c2 = st.columns([6, 1])
        ref["sitasi"] = c1.text_input(f"Referensi #{idx + 1}", ref["sitasi"], key=mk_key(f"ref_{idx}"))
        if c2.button("Hapus", key=mk_key(f"ref_del_{idx}")):
            st.session_state.referensi_data.pop(idx)
            st.rerun()
    if st.button("➕ Tambah Referensi", key=mk_key("tambah_referensi")):
        st.session_state.referensi_data.append({"sitasi": ""})
        st.rerun()

# --- Tab: Komponen Penilaian (radio per CPMK — bebas isu 2x klik data_editor) ---
with tab_nilai:
    st.subheader("Komponen Penilaian")
    st.caption(
        "Pilih satu kategori penilaian untuk tiap CPMK. Satu kategori boleh dipakai oleh "
        "lebih dari satu CPMK, tapi satu CPMK hanya boleh masuk satu kategori."
    )
    for i in range(1, 6):
        current = st.session_state.komponen_data.get(i)
        idx = KATEGORI_PENILAIAN.index(current) if current in KATEGORI_PENILAIAN else 0
        choice = st.radio(
            f"CPMK-{i}", KATEGORI_PENILAIAN, index=idx, horizontal=True,
            key=mk_key(f"komp_radio_{i}"),
        )
        st.session_state.komponen_data[i] = choice

    st.divider()
    st.markdown("**Bobot Komponen Penilaian (baku):**")
    cols = st.columns(4)
    for idx, (k, v) in enumerate(BOBOT_KATEGORI.items()):
        cols[idx].metric(k, f"{v}%")

# --- Tab: Pratinjau & Ekspor ---
with tab_export:
    st.subheader("Pratinjau")
    ready = len(st.session_state.cpl_selected) == N_CPL_WAJIB
    if not ready:
        st.info("Lengkapi pemilihan 5 CPL di tab 'CPL & CPMK' terlebih dahulu untuk mengaktifkan ekspor.")
    else:
        st.markdown(f"### {mk_row['Nama Mata Kuliah']} ({mk_row['Kode MK']})")
        st.write(f"**Prodi:** {st.session_state.prodi_sel} · **SKS:** {mk_row['SKS']} · **Semester:** {mk_row['Semester']}")
        st.dataframe(pd.DataFrame([
            {"CPMK": f"CPMK-{i}",
             "Deskripsi (dengan kode CPL)": with_code(
                 st.session_state.cpmk_data[i]["deskripsi"], st.session_state.cpmk_data[i]["cpl_kode"])[:100]}
            for i in range(1, 6)
        ]), use_container_width=True, hide_index=True)

        st.divider()
        st.subheader("Validasi Dokumen")
        st.caption(
            "Nama-nama ini akan muncul di blok tanda tangan pada bagian akhir dokumen (tanpa QR code — "
            "ruang kosong disediakan untuk tanda tangan fisik)."
        )
        info = st.session_state.info_umum
        info["tanggal_dokumen"] = st.text_input(
            "Tanggal Dokumen (mis. 28/10/2025)", info["tanggal_dokumen"], key=mk_key("tanggal_dokumen"),
        )
        vc1, vc2, vc3 = st.columns(3)
        info["nama_kaprodi"] = vc1.text_input("Nama Ketua Prodi", info["nama_kaprodi"], key=mk_key("nama_kaprodi"))
        info["nama_koordinator"] = vc2.text_input(
            "Nama Koordinator MK/Bidang Keahlian", info["nama_koordinator"], key=mk_key("nama_koordinator"),
        )
        info["nama_penyusun"] = vc3.text_input(
            "Nama Dosen Penyusun", info["nama_penyusun"], key=mk_key("nama_penyusun"),
        )
        info["nama_biro_pjm"] = st.text_input(
            "Nama Kepala Biro Penjaminan Mutu", info["nama_biro_pjm"], key=mk_key("nama_biro_pjm"),
        )

        st.divider()
        col1, col2 = st.columns(2)
        with col1:
            excel_buf = export_to_excel(mk_row, cpl_df)
            st.download_button("⬇️ Unduh Excel", data=excel_buf,
                                file_name=f"RPS_{mk_row['Kode MK']}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="download_excel_btn")
        with col2:
            pdf_buf = build_pdf(
                prodi=st.session_state.prodi_sel, mk_row=mk_row, cpl_df=cpl_df,
                info_umum=st.session_state.info_umum, cpmk_data=st.session_state.cpmk_data,
                pertemuan_data=st.session_state.pertemuan_data,
                referensi_data=st.session_state.referensi_data,
                komponen_data=st.session_state.komponen_data,
                bobot_kategori=BOBOT_KATEGORI,
            )
            st.download_button("⬇️ Unduh PDF", data=pdf_buf,
                                file_name=f"RPS_{mk_row['Kode MK']}.pdf", mime="application/pdf",
                                key="download_pdf_btn")
